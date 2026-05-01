"""Plan 6 Task 7：生成型 tool 测试（≥10 case）。"""
from __future__ import annotations

import base64
import io
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

# ===== helpers =====

def _make_template_docx(text: str = "客户：{{customer_name}}") -> bytes:
    """创建最小 valid docx 字节（用 python-docx）。"""
    from docx import Document
    doc = Document()
    doc.add_paragraph(text)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _make_template_docx_b64(text: str = "客户：{{customer_name}}") -> str:
    """返回 base64 编码的 docx bytes（ContractTemplate.file_storage_key 格式）。"""
    return base64.b64encode(_make_template_docx(text)).decode()


# ===== fixtures =====

@pytest.fixture
def mock_sender():
    """注入 mock DingTalkSender 到 generate_tools 模块。"""
    from hub.agent.tools import generate_tools
    s = MagicMock()
    s.send_file = AsyncMock()
    erp = MagicMock()
    # I3: 改用精确 get_customer 而非 search_customers keyword 搜索
    erp.get_customer = AsyncMock(return_value={"id": 100, "name": "测试客户", "address": "上海市"})
    # v2 staging review #5: generate_contract_draft 渲染前调 get_account_set 拉甲方
    erp.get_account_set = AsyncMock(return_value={
        "id": 1, "name": "广州启领信息科技有限公司",
        "company_name": "广州启领信息科技有限公司",
        "bank_name": "中国建设银行股份有限公司广州江湾路支行",
        "bank_account": "44050146081300000741",
        "tax_id": "",
    })
    generate_tools.set_dependencies(sender=s, erp=erp)
    yield s
    generate_tools.set_dependencies(sender=None, erp=None)


SAMPLE_ITEMS = [
    {"name": "讯飞x5", "qty": 50, "price": 1000.0, "subtotal": 50000.0},
]


# ===== ExcelExporter tests =====

async def test_excel_exporter_handles_nested_dict_value():
    """ExcelExporter：dict 类型 cell 不抛 openpyxl ValueError，转成 JSON 字符串。"""
    from openpyxl import load_workbook

    from hub.agent.document.excel import ExcelExporter

    data = [{"名称": "A", "extra": {"nested": "value", "num": 1}}]
    exporter = ExcelExporter()
    result = await exporter.export(table_data=data)

    assert result[:2] == b"PK", "应是合法 xlsx"
    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    # 找 extra 列
    headers = [ws.cell(1, i + 1).value for i in range(ws.max_column)]
    extra_col = headers.index("extra") + 1
    cell_val = ws.cell(2, extra_col).value
    assert isinstance(cell_val, str), "dict 应被序列化为 str"
    assert "nested" in cell_val


async def test_excel_exporter_handles_list_value():
    """ExcelExporter：list 类型 cell 不抛 openpyxl ValueError，转成 JSON 字符串。"""
    from openpyxl import load_workbook

    from hub.agent.document.excel import ExcelExporter

    data = [{"名称": "B", "tags": ["tag1", "tag2"]}]
    exporter = ExcelExporter()
    result = await exporter.export(table_data=data)

    assert result[:2] == b"PK"
    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = [ws.cell(1, i + 1).value for i in range(ws.max_column)]
    tags_col = headers.index("tags") + 1
    cell_val = ws.cell(2, tags_col).value
    assert isinstance(cell_val, str), "list 应被序列化为 str"
    assert "tag1" in cell_val


async def test_export_to_excel_basic():
    """export_to_excel 生成合法 xlsx（以 PK magic bytes 开头）。"""
    from hub.agent.document.excel import ExcelExporter

    data = [{"名称": "商品A", "数量": 10}, {"名称": "商品B", "数量": 20}]
    exporter = ExcelExporter()
    result = await exporter.export(table_data=data)

    assert isinstance(result, bytes)
    assert len(result) > 0
    # xlsx 是 ZIP 格式，magic bytes = PK
    assert result[:2] == b"PK", "xlsx 应以 PK（ZIP magic bytes）开头"


async def test_export_to_excel_empty_table_data():
    """空 list 也能生成合法 xlsx（不崩溃）。"""
    from hub.agent.document.excel import ExcelExporter

    exporter = ExcelExporter()
    result = await exporter.export(table_data=[])

    assert isinstance(result, bytes)
    assert result[:2] == b"PK"


async def test_export_to_excel_headers_from_first_row():
    """表头取各行 key 并集，保留首行顺序。"""
    from openpyxl import load_workbook

    from hub.agent.document.excel import ExcelExporter

    data = [
        {"姓名": "张三", "年龄": 30},
        {"姓名": "李四", "年龄": 25, "部门": "销售"},
    ]
    exporter = ExcelExporter()
    result = await exporter.export(table_data=data)

    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = [ws.cell(1, i + 1).value for i in range(ws.max_column)]
    assert "姓名" in headers
    assert "年龄" in headers
    assert "部门" in headers


# ===== export_to_excel tool tests =====

async def test_export_to_excel_tool_no_binding(mock_sender):
    """用户无 active 钉钉绑定 → file_sent=False + warning。"""
    from hub.agent.tools.generate_tools import export_to_excel

    # ChannelUserBinding.filter().first() → None
    with patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding:
        qs = MagicMock()
        qs.first = AsyncMock(return_value=None)
        mock_binding.filter.return_value = qs

        result = await export_to_excel(
            table_data=[{"a": 1}],
            file_name="test",
            hub_user_id=1,
            conversation_id="conv-1",
            acting_as_user_id=1,
        )

    assert result["file_sent"] is False
    assert "warning" in result
    mock_sender.send_file.assert_not_called()


async def test_export_to_excel_tool_basic(mock_sender):
    """export_to_excel 有绑定时调 send_file，返 file_sent=True + rows_count。"""
    from hub.agent.tools.generate_tools import export_to_excel

    binding = MagicMock()
    binding.channel_userid = "ding-u1"

    with patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding:
        qs = MagicMock()
        qs.first = AsyncMock(return_value=binding)
        mock_binding.filter.return_value = qs

        result = await export_to_excel(
            table_data=[{"商品": "A", "数量": 1}, {"商品": "B", "数量": 2}],
            file_name="export",
            hub_user_id=1,
            conversation_id="conv-1",
            acting_as_user_id=1,
        )

    assert result["file_sent"] is True
    assert result["rows_count"] == 2
    assert result["file_name"].endswith(".xlsx")
    mock_sender.send_file.assert_awaited_once()


async def test_export_to_excel_send_file_error_propagates(mock_sender):
    """send_file 抛 DingTalkSendError → export_to_excel 透传异常。"""
    from hub.adapters.channel.dingtalk_sender import DingTalkSendError
    from hub.agent.tools.generate_tools import export_to_excel

    mock_sender.send_file = AsyncMock(side_effect=DingTalkSendError("网络超时"))

    binding = MagicMock()
    binding.channel_userid = "ding-u1"

    with patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding:
        qs = MagicMock()
        qs.first = AsyncMock(return_value=binding)
        mock_binding.filter.return_value = qs

        with pytest.raises(DingTalkSendError, match="网络超时"):
            await export_to_excel(
                table_data=[{"a": 1}],
                file_name="err_test.xlsx",
                hub_user_id=1,
                conversation_id="conv-1",
                acting_as_user_id=1,
            )


# ===== generate_contract_draft tests =====

async def test_generate_contract_draft_basic(mock_sender):
    """generate_contract_draft：有合法模板 + 有 binding → 创建 ContractDraft + file_sent=True。"""
    from hub.agent.tools.generate_tools import generate_contract_draft

    # mock ContractTemplate
    template = MagicMock()
    template.id = 1
    template.is_active = True
    template.file_storage_key = _make_template_docx_b64()
    template.placeholders = []

    # mock ContractDraft.create → 假 draft
    draft = MagicMock()
    draft.id = 42
    draft.save = AsyncMock()

    # mock binding
    binding = MagicMock()
    binding.channel_userid = "ding-u1"

    with (
        # ContractRenderer 内部用的 ContractTemplate（hub.agent.document.contract 模块）
        patch("hub.agent.document.contract.ContractTemplate") as mock_contract_template,
        patch("hub.agent.tools.generate_tools.ContractDraft") as mock_draft,
        patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding,
    ):
        # ContractTemplate.filter().first()
        tqs = MagicMock()
        tqs.first = AsyncMock(return_value=template)
        mock_contract_template.filter.return_value = tqs

        # ContractDraft.create()
        mock_draft.create = AsyncMock(return_value=draft)
        # v8 review #17：fingerprint 幂等查询 mock（默认 .first() 返 None 走 create 分支）
        mock_draft.filter.return_value.first = AsyncMock(return_value=None)

        # ChannelUserBinding.filter().first()
        bqs = MagicMock()
        bqs.first = AsyncMock(return_value=binding)
        mock_binding.filter.return_value = bqs

        result = await generate_contract_draft(
            template_id=1,
            customer_id=100,
            items=SAMPLE_ITEMS,
            hub_user_id=1,
            conversation_id="conv-1",
            acting_as_user_id=1,
        )

    assert result["draft_id"] == 42
    assert result["file_sent"] is True
    assert result["file_name"].endswith(".docx")
    mock_draft.create.assert_awaited_once()
    # C1 验证：rendered_file_storage_key 传 None
    create_kwargs = mock_draft.create.call_args.kwargs
    assert create_kwargs["rendered_file_storage_key"] is None
    mock_sender.send_file.assert_awaited_once()
    # M4 验证：send_file 成功后 status 推进到 sent
    draft.save.assert_awaited_once()


async def test_generate_contract_draft_template_not_found(mock_sender):
    """template_id 不存在 → 返回 file_sent=False + error 字段（I4：转友好返回 dict）。"""
    from hub.agent.tools.generate_tools import generate_contract_draft

    with patch("hub.agent.document.contract.ContractTemplate") as mock_template:
        tqs = MagicMock()
        tqs.first = AsyncMock(return_value=None)
        mock_template.filter.return_value = tqs

        result = await generate_contract_draft(
            template_id=999,
            customer_id=100,
            items=SAMPLE_ITEMS,
            hub_user_id=1,
            conversation_id="conv-1",
            acting_as_user_id=1,
        )

    assert result["file_sent"] is False
    assert result["draft_id"] is None
    assert "error" in result
    assert "999" in result["error"]


# ===== v8 staging review #17: fingerprint 幂等覆盖 items + extras =====

def test_compute_contract_fingerprint_stable():
    """同输入 → 同 fingerprint（dict key 顺序不影响）。"""
    from hub.agent.tools.generate_tools import _compute_contract_fingerprint

    fp1 = _compute_contract_fingerprint(
        template_id=1, customer_id=11,
        items=[{"product_id": 5030, "qty": 10, "price": 4000}],
        extras={"contract_no": "C-001", "payment_terms": "30 天"},
    )
    fp2 = _compute_contract_fingerprint(
        template_id=1, customer_id=11,
        # 同 items + extras 但 dict key 顺序倒过来
        items=[{"price": 4000, "product_id": 5030, "qty": 10}],
        extras={"payment_terms": "30 天", "contract_no": "C-001"},
    )
    assert fp1 == fp2
    assert len(fp1) == 64  # sha256 hex


def test_compute_contract_fingerprint_extras_change_creates_new():
    """改 extras 任一字段 → fingerprint 变（不会复用旧 draft）。

    v8 review #17 防的就是：用户改了合同号 / 付款条款 / 收货地址等，
    新文件已渲染发出，但代码 bug 复用旧 draft DB 记录 → 审计失真。
    """
    from hub.agent.tools.generate_tools import _compute_contract_fingerprint

    base = dict(
        template_id=1, customer_id=11,
        items=[{"product_id": 5030, "qty": 10, "price": 4000}],
    )
    fp_a = _compute_contract_fingerprint(**base, extras={"contract_no": "C-001"})
    # 只改了合同号
    fp_b = _compute_contract_fingerprint(**base, extras={"contract_no": "C-002"})
    assert fp_a != fp_b

    # 加了一个 payment_terms 字段
    fp_c = _compute_contract_fingerprint(
        **base, extras={"contract_no": "C-001", "payment_terms": "30 天"},
    )
    assert fp_a != fp_c

    # 改了 items 数量
    fp_d = _compute_contract_fingerprint(
        template_id=1, customer_id=11,
        items=[{"product_id": 5030, "qty": 20, "price": 4000}],  # 10 → 20
        extras={"contract_no": "C-001"},
    )
    assert fp_a != fp_d


async def test_generate_contract_draft_no_active_binding(mock_sender):
    """用户无 active 钉钉绑定 → 草稿仍持久化 + file_sent=False。"""
    from hub.agent.tools.generate_tools import generate_contract_draft

    template = MagicMock()
    template.id = 1
    template.is_active = True
    template.file_storage_key = _make_template_docx_b64()
    template.placeholders = []

    draft = MagicMock()
    draft.id = 43

    with (
        patch("hub.agent.document.contract.ContractTemplate") as mock_contract_template,
        patch("hub.agent.tools.generate_tools.ContractDraft") as mock_draft,
        patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding,
    ):
        tqs = MagicMock()
        tqs.first = AsyncMock(return_value=template)
        mock_contract_template.filter.return_value = tqs

        mock_draft.create = AsyncMock(return_value=draft)
        # v8 review #17：fingerprint 幂等查询 mock（默认 .first() 返 None 走 create 分支）
        mock_draft.filter.return_value.first = AsyncMock(return_value=None)

        bqs = MagicMock()
        bqs.first = AsyncMock(return_value=None)  # 没有 binding
        mock_binding.filter.return_value = bqs

        result = await generate_contract_draft(
            template_id=1,
            customer_id=100,
            items=SAMPLE_ITEMS,
            hub_user_id=1,
            conversation_id="conv-1",
            acting_as_user_id=1,
        )

    assert result["file_sent"] is False
    assert "warning" in result
    assert result["draft_id"] == 43  # 草稿已持久化
    mock_draft.create.assert_awaited_once()
    mock_sender.send_file.assert_not_called()
    # C2 验证：ChannelUserBinding.filter 必须带 channel_type="dingtalk"
    filter_kwargs = mock_binding.filter.call_args.kwargs
    assert filter_kwargs.get("channel_type") == "dingtalk"


async def test_generate_contract_draft_send_file_failure_propagates(mock_sender):
    """send_file 抛错 → 草稿已存 + 异常透传（让 worker 转死信重试）。"""
    from hub.adapters.channel.dingtalk_sender import DingTalkSendError
    from hub.agent.tools.generate_tools import generate_contract_draft

    mock_sender.send_file = AsyncMock(side_effect=DingTalkSendError("钉钉超时"))

    template = MagicMock()
    template.id = 1
    template.is_active = True
    template.file_storage_key = _make_template_docx_b64()
    template.placeholders = []

    draft = MagicMock()
    draft.id = 44

    binding = MagicMock()
    binding.channel_userid = "ding-u1"

    with (
        patch("hub.agent.document.contract.ContractTemplate") as mock_contract_template,
        patch("hub.agent.tools.generate_tools.ContractDraft") as mock_draft,
        patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding,
    ):
        tqs = MagicMock()
        tqs.first = AsyncMock(return_value=template)
        mock_contract_template.filter.return_value = tqs

        mock_draft.create = AsyncMock(return_value=draft)
        # v8 review #17：fingerprint 幂等查询 mock（默认 .first() 返 None 走 create 分支）
        mock_draft.filter.return_value.first = AsyncMock(return_value=None)

        bqs = MagicMock()
        bqs.first = AsyncMock(return_value=binding)
        mock_binding.filter.return_value = bqs

        with pytest.raises(DingTalkSendError, match="钉钉超时"):
            await generate_contract_draft(
                template_id=1,
                customer_id=100,
                items=SAMPLE_ITEMS,
                hub_user_id=1,
                conversation_id="conv-1",
                acting_as_user_id=1,
            )

    # 草稿已创建（在 send_file 之前）
    mock_draft.create.assert_awaited_once()


async def test_generate_contract_draft_get_customer_failed_returns_error(mock_sender):
    """v8 review #12：get_customer 失败时不再 fallback 占位，返 error 让 LLM 重调
    （旧行为是 fallback 到"客户N"占位继续生成 → 用户收到错的合同）。"""
    from hub.adapters.downstream.erp4 import ErpNotFoundError
    from hub.agent.tools import generate_tools
    from hub.agent.tools.generate_tools import generate_contract_draft

    # 覆盖 erp.get_customer 为抛 ErpNotFoundError
    erp = MagicMock()
    erp.get_customer = AsyncMock(side_effect=ErpNotFoundError("customer 100 not found"))
    # v2 staging review #5: 加 get_account_set mock（generate_contract_draft 渲染前必调）
    erp.get_account_set = AsyncMock(return_value={
        "id": 1, "name": "测试启领", "company_name": "测试启领",
        "bank_name": "测试银行", "bank_account": "001",
    })
    generate_tools.set_dependencies(sender=generate_tools._dingtalk_sender, erp=erp)

    template = MagicMock()
    template.id = 1
    template.is_active = True
    template.file_storage_key = _make_template_docx_b64()
    template.placeholders = []

    draft = MagicMock()
    draft.id = 55
    draft.save = AsyncMock()

    binding = MagicMock()
    binding.channel_userid = "ding-u1"

    with (
        patch("hub.agent.document.contract.ContractTemplate") as mock_contract_template,
        patch("hub.agent.tools.generate_tools.ContractDraft") as mock_draft,
        patch("hub.agent.tools.generate_tools.ChannelUserBinding") as mock_binding,
    ):
        tqs = MagicMock()
        tqs.first = AsyncMock(return_value=template)
        mock_contract_template.filter.return_value = tqs
        mock_draft.create = AsyncMock(return_value=draft)
        # v8 review #17：fingerprint 幂等查询 mock（默认 .first() 返 None 走 create 分支）
        mock_draft.filter.return_value.first = AsyncMock(return_value=None)
        bqs = MagicMock()
        bqs.first = AsyncMock(return_value=binding)
        mock_binding.filter.return_value = bqs

        result = await generate_contract_draft(
            template_id=1,
            customer_id=100,
            items=SAMPLE_ITEMS,
            hub_user_id=1,
            conversation_id="conv-fallback",
            acting_as_user_id=1,
        )

    # 客户不存在 → 返 error 不生成合同（v8 review #12 改）
    assert result["file_sent"] is False
    assert result["draft_id"] is None
    assert "ID 100" in result["error"]
    assert "search_customers" in result["error"]
    # 恢复 mock_sender erp 注入
    generate_tools.set_dependencies(sender=generate_tools._dingtalk_sender, erp=None)


# ===== generate_price_quote tests =====

async def test_generate_price_quote_no_template(mock_sender):
    """未配置报价模板 → 返回 file_sent=False + error 字段。"""
    from hub.agent.tools.generate_tools import generate_price_quote

    with patch("hub.agent.tools.generate_tools.ContractTemplate") as mock_template:
        qs = MagicMock()
        qs.first = AsyncMock(return_value=None)
        mock_template.filter.return_value = qs

        result = await generate_price_quote(
            customer_id=100,
            items=SAMPLE_ITEMS,
            hub_user_id=1,
            conversation_id="conv-1",
            acting_as_user_id=1,
        )

    assert result["file_sent"] is False
    assert "error" in result
    assert result["draft_id"] is None


# ===== register_all test =====

async def test_register_all_registers_3_generate_tools():
    """register_all 成功注册 3 个 GENERATE 类 tool，不抛 ToolRegistrationError。"""
    from unittest.mock import MagicMock

    from hub.agent.tools.generate_tools import register_all
    from hub.agent.tools.types import ToolType

    registry = MagicMock()
    register_all(registry)

    assert registry.register.call_count == 3

    # 验证每个 tool 的 tool_type = GENERATE
    call_kwargs = [call.kwargs for call in registry.register.call_args_list]
    for kw in call_kwargs:
        assert kw["tool_type"] == ToolType.GENERATE


# ===== DocumentStorage tests =====

async def test_document_storage_encrypt_decrypt_roundtrip():
    """DocumentStorage put (encrypt=True) → get (decrypt=True) 还原原始 bytes。"""
    from hub.agent.document.storage import DocumentStorage

    with patch("hub.agent.document.storage.get_settings") as mock_settings:
        settings = MagicMock()
        settings.master_key_bytes = bytes(range(32))
        mock_settings.return_value = settings

        storage = DocumentStorage()
        original = b"PK\x03\x04docx-content-bytes"  # mock xlsx/docx bytes
        encrypted = await storage.put(original, encrypted=True)

        assert encrypted != original, "加密后 bytes 应与原文不同"

        restored = await storage.get(encrypted, encrypted=True)
        assert restored == original, "解密后应还原原始 bytes"


async def test_document_storage_no_encrypt():
    """encrypted=False 时 put/get 直接返回原始 bytes（不加密）。"""
    from hub.agent.document.storage import DocumentStorage

    storage = DocumentStorage()
    original = b"hello world"
    stored = await storage.put(original, encrypted=False)
    assert stored == original

    recovered = await storage.get(stored, encrypted=False)
    assert recovered == original
